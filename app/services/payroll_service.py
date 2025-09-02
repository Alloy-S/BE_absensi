from app.execption.custom_execption import GeneralException, GeneralExceptionWithParam
from app.repositories.grup_gaji_repository import GrupGajiRepository
from app.repositories.kode_perhitungan_repository import KodePerhitunganRepository
from app.repositories.laporan_repository import LaporanRepository
from app.repositories.riwayat_penggajian_repository import RiwayatPenggajianRepository
from app.repositories.user_repository import UserRepository
from app.utils.app_constans import AppConstants
from app.utils.error_code import ErrorCode
from app.database import db
from decimal import Decimal, ROUND_HALF_UP
import pandas as pd
import io
import base64
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side



class PayrollService:
    @staticmethod
    def proses_gaji_grup(username, validated):

        user = UserRepository.get_user_by_username(username)

        if not user:
            raise GeneralExceptionWithParam(ErrorCode.RESOURCE_NOT_FOUND,
                                            params={"resource": AppConstants.USER_RESOURCE.value})

        grup_gaji_id = validated.get('grup_gaji_id')
        periode_start = validated.get('periode_start')
        periode_end = validated.get('periode_end')

        data_karyawan = GrupGajiRepository.get_grup_gaji_users(grup_gaji_id)

        aturan_gaji = GrupGajiRepository.get_grup_gaji_detail_by_id(grup_gaji_id)

        list_user_id = [karyawan.id for karyawan in data_karyawan]

        data_laporan = LaporanRepository.generate_laporan_periode(list_user_id, periode_start, periode_end)

        kode_field_laporan = KodePerhitunganRepository.get_all_kode_perhitungan()

        kode_field_laporan = {
            item.kode: item.field_laporan for item in kode_field_laporan
        }

        hasil_penggajian_final = []
        total_gaji_keseluruhan = Decimal('0.0')

        for karyawan in data_karyawan:
            total_tunjangan = Decimal('0.0')
            total_potongan = Decimal('0.0')

            rincian_perhitungan = []
            data_laporan_karyawan = data_laporan[karyawan.id]

            if not data_laporan_karyawan:
                continue
            print(data_laporan_karyawan)

            for aturan in aturan_gaji:
                hasil_komponen = Decimal(0.0)
                print(aturan)
                operasi = 'x'
                nilai = Decimal(1.0)

                nilai_uang_final = aturan.nilai_statis

                if aturan.use_nilai_dinamis:
                    nilai_uang_final = PayrollService.get_data_komponen(kode_field_laporan, aturan.kode_nilai_dinamis,
                                                                     data_laporan_karyawan)

                if aturan.use_kondisi:
                    nilai_kondisi = PayrollService.get_data_komponen(kode_field_laporan, aturan.kode_kondisi,
                                                                     data_laporan_karyawan)

                    if aturan.min_kondisi <= nilai_kondisi <= aturan.max_kondisi:
                        hasil_komponen = PayrollService.perhitungan_operasi(nilai, operasi, nilai_uang_final)
                    else:
                        continue

                elif aturan.use_formula:
                    nilai = PayrollService.get_data_komponen(kode_field_laporan, aturan.kode_formula,
                                                             data_laporan_karyawan)

                    operasi = aturan.operation_sum
                    hasil_komponen = PayrollService.perhitungan_operasi(nilai, operasi,
                                                                        nilai_uang_final)

                else:
                    hasil_komponen = PayrollService.perhitungan_operasi(nilai, operasi,
                                                                        nilai_uang_final)

                if aturan.tipe == 'TUNJANGAN':
                    total_tunjangan += hasil_komponen
                else:
                    total_potongan += hasil_komponen

                rincian_perhitungan.append({
                    'komponen': aturan.kom_name,
                    'tipe': aturan.tipe,
                    'jumlah': hasil_komponen,
                    'nilai_a': nilai_uang_final,
                    'nilai_b': nilai,
                    'operasi': operasi,
                })

            gaji = total_tunjangan - total_potongan
            total_gaji_keseluruhan += gaji

            hasil_penggajian_final.append({
                'user_id': karyawan.id,
                'nama_karyawan': karyawan.fullname,
                'total_tunjangan': total_tunjangan,
                'total_potongan': total_potongan,
                'gaji': gaji,
                'rincian': rincian_perhitungan,
            })

        riwayat_data_to_save = {
            "grup_gaji_id": grup_gaji_id,
            "periode_start": periode_start,
            "periode_end": periode_end,
            "total_karyawan": len(hasil_penggajian_final),
            "total_gaji_keseluruhan": total_gaji_keseluruhan.quantize(Decimal('0.01'), ROUND_HALF_UP),
            "created_by": user.id,
            "hasil_karyawan": hasil_penggajian_final
        }

        riwayat_id = RiwayatPenggajianRepository.create_draft_riwayat_penggajian(riwayat_data_to_save)

        return {
            'riwayat_id': riwayat_id,
            'data': hasil_penggajian_final
        }

    @staticmethod
    def perhitungan_operasi(nilai_a, operasi, nilai_b):

        nilai_a_decimal = Decimal(nilai_a)

        if operasi == '/':
            nilai_final = nilai_a_decimal / nilai_b
        elif operasi == 'x':
            nilai_final = nilai_a_decimal * nilai_b
        elif operasi == '+':
            nilai_final = nilai_a_decimal + nilai_b
        elif operasi == '-':
            nilai_final = nilai_a_decimal - nilai_b
        else:
            raise GeneralExceptionWithParam(ErrorCode.RESOURCE_NOT_FOUND, params={'resource': 'Tanda Operasi'})

        return nilai_final

    @staticmethod
    def get_data_komponen(list_kode, kode_komponen, data_laporan):
        return data_laporan[list_kode[kode_komponen]]

    @staticmethod
    def finalisasi_riwayat_penggajian(riwayat_id):
        riwayat = RiwayatPenggajianRepository.get_riwayat_penggajian_by_id(riwayat_id)

        if not riwayat:
            raise GeneralExceptionWithParam(ErrorCode.RESOURCE_NOT_FOUND,
                                            params={'resource': AppConstants.RIWAYAT_PENGGAJIAN_RESOURCE.value})

        if riwayat.status == 'FINAL':
            raise GeneralException(ErrorCode.ALREADY_FINAL_STATUS)

        RiwayatPenggajianRepository.finalisasi_riwayat_penggajian(riwayat)

    @staticmethod
    def get_riwayat_by_id(riwayat_id):
        riwayat = RiwayatPenggajianRepository.get_riwayat_penggajian_by_id(riwayat_id)

        if not riwayat:
            raise GeneralExceptionWithParam(ErrorCode.RESOURCE_NOT_FOUND,
                                            params={'resource': AppConstants.RIWAYAT_PENGGAJIAN_RESOURCE.value})

        return riwayat

    @staticmethod
    def get_riwayat_penggajian_pagination(request):
        return RiwayatPenggajianRepository.get_riwayat_penggajian_pagination(request.get('page'), request.get('size'),
                                                                             request.get('periode_start'),
                                                                             request.get('periode_end'),
                                                                             request.get('grup_gaji'),
                                                                             request.get('status'))

    @staticmethod
    def export_riwayat_penggajian_to_excel(riwayat_id):
        riwayat = RiwayatPenggajianRepository.get_riwayat_penggajian_by_id(riwayat_id)

        if not riwayat:
            raise GeneralExceptionWithParam(ErrorCode.RESOURCE_NOT_FOUND,
                                            params={'resource': AppConstants.RIWAYAT_PENGGAJIAN_RESOURCE.value})

        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:

            bold_font = Font(bold=True)
            currency_format = '"Rp"#,##0'

            info_data = {
                "Keterangan": ["Judul Laporan", "Grup Gaji", "Periode", "Status", "Tanggal Dibuat", "Dibuat Oleh"],
                "Detail": [
                    "Laporan Detail Penggajian",
                    riwayat.grup_gaji.grup_name,
                    f"{riwayat.periode_start.strftime('%d %B %Y')} - {riwayat.periode_end.strftime('%d %B %Y')}",
                    riwayat.status,
                    riwayat.date_created.strftime('%d %B %Y %H:%M:%S'),
                    riwayat.user.fullname
                ]
            }
            df_info = pd.DataFrame(info_data)
            df_info.to_excel(writer, sheet_name="Informasi Laporan", index=False, header=False)


            ringkasan_list = []
            for detail in riwayat.riwayat_penggajian_detail:
                ringkasan_list.append({
                    "NIP": detail.user.data_karyawan.nip,
                    "Nama Karyawan": detail.user.fullname,
                    "Total Tunjangan": detail.total_tunjangan,
                    "Total Potongan": detail.total_potongan,
                    "Gaji": detail.gaji
                })

            df_ringkasan = pd.DataFrame(ringkasan_list)

            total = df_ringkasan[["Total Tunjangan", "Total Potongan", "Gaji"]].sum().to_frame().T
            total["Nama Karyawan"] = "GRAND TOTAL"
            df_ringkasan = pd.concat([df_ringkasan, total], ignore_index=True)

            df_ringkasan.to_excel(writer, sheet_name="Ringkasan Gaji", index_label="No", startrow=1)

            ws_ringkasan = writer.sheets['Ringkasan Gaji']

            for cell in ws_ringkasan['D']:
                if cell.row > 2:
                    cell.number_format = currency_format

            for cell in ws_ringkasan['E']:
                if cell.row > 2:
                    cell.number_format = currency_format

            for cell in ws_ringkasan['F']:
                if cell.row > 2:
                    cell.number_format = currency_format

            ws = writer.book.create_sheet("Detail per Karyawan")

            current_row = 1

            for detail in riwayat.riwayat_penggajian_detail:
                ws.cell(row=current_row, column=1, value="Nama Karyawan:").font = bold_font
                ws.cell(row=current_row, column=2, value=detail.user.fullname)
                ws.cell(row=current_row, column=3, value="NIP:").font = bold_font
                ws.cell(row=current_row, column=4, value=detail.user.data_karyawan.nip)
                current_row += 2

                ws.cell(row=current_row, column=1, value="Tunjangan").font = bold_font
                ws.cell(row=current_row, column=7, value="Jumlah").font = bold_font
                current_row += 1

                for rincian in detail.riwayat_penggajian_rincian:
                    if rincian.tipe == 'TUNJANGAN':
                        ws.cell(row=current_row, column=2, value=rincian.komponen)
                        ws.cell(row=current_row, column=3, value=rincian.nilai_a)
                        ws.cell(row=current_row, column=4, value=rincian.operasi)
                        ws.cell(row=current_row, column=5, value=rincian.nilai_b)
                        cell = ws.cell(row=current_row, column=7, value=rincian.jumlah)
                        cell.number_format = currency_format
                        current_row += 1

                ws.cell(row=current_row, column=6, value="Total Tunjangan").font = bold_font
                cell = ws.cell(row=current_row, column=7, value=detail.total_tunjangan)
                cell.font = bold_font
                cell.number_format = currency_format
                current_row += 2

                ws.cell(row=current_row, column=1, value="Potongan").font = bold_font
                ws.cell(row=current_row, column=7, value="Jumlah").font = bold_font
                current_row += 1

                for rincian in detail.riwayat_penggajian_rincian:
                    if rincian.tipe == 'POTONGAN':
                        ws.cell(row=current_row, column=2, value=rincian.komponen)
                        ws.cell(row=current_row, column=3, value=rincian.nilai_a)
                        ws.cell(row=current_row, column=4, value=rincian.operasi)
                        ws.cell(row=current_row, column=5, value=rincian.nilai_b)
                        cell = ws.cell(row=current_row, column=7, value=rincian.jumlah)

                        cell.number_format = currency_format
                        current_row += 1

                ws.cell(row=current_row, column=6, value="Total Potongan").font = bold_font
                cell = ws.cell(row=current_row, column=7, value=detail.total_potongan)
                cell.font = bold_font
                cell.number_format = currency_format
                current_row += 2

                ws.cell(row=current_row, column=6, value="GAJI").font = Font(bold=True, size=12)
                cell = ws.cell(row=current_row, column=7, value=detail.gaji)
                cell.font = Font(bold=True, size=12)
                cell.number_format = currency_format
                current_row += 3

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 25
            ws.column_dimensions['G'].width = 25



        output.seek(0)

        excel_base64 = base64.b64encode(output.getvalue()).decode('utf-8')

        file_name = f'export_gaji_{riwayat.periode_start}_-_{riwayat.periode_end}.xlsx'

        return {
            'filename': file_name,
            'file': excel_base64
        }
